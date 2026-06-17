"""
Módulo de Abas Auxiliares para o Conversor DOit
Gera abas extras no Excel final: Dados Bancários, Plano de Contas, Pendências.
Extensível para futuras abas (Contatos, Responsáveis, Documentos, etc.)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import re


# ============================================================
# DETECÇÃO DE DADOS BANCÁRIOS
# ============================================================

CAMPOS_BANCARIOS = {
    'favorecido': ['favorecido', 'razão social', 'razao social', 'titular', 'nome', 'empresa'],
    'cnpj': ['cnpj', 'cpf/cnpj', 'cpf', 'documento'],
    'banco': ['banco', 'instituição', 'instituicao', 'inst. financeira'],
    'agencia': ['agência', 'agencia', 'ag'],
    'conta': ['conta', 'conta corrente', 'c/c', 'nº conta', 'numero conta'],
    'tipo_conta': ['tipo de conta', 'tipo conta', 'modalidade'],
    'pix': ['pix', 'chave pix', 'chave', 'chavepix'],
}


def detectar_dados_bancarios(df_entrada: pd.DataFrame, todas_abas: dict = None) -> dict:
    """
    Detecta dados bancários na planilha de entrada ou em abas específicas.
    
    Args:
        df_entrada: DataFrame principal
        todas_abas: dict com {nome_aba: DataFrame} para buscar em múltiplas abas
    
    Returns:
        dict com dados bancários encontrados ou None se não encontrar
    """
    dados = {}
    
    # Fontes para buscar
    fontes = {'principal': df_entrada}
    if todas_abas:
        fontes.update(todas_abas)
    
    for nome_fonte, df in fontes.items():
        if df is None or df.empty:
            continue
        
        colunas_lower = {str(c).lower().strip(): c for c in df.columns}
        
        # Tentar encontrar por colunas
        for campo_padrao, variantes in CAMPOS_BANCARIOS.items():
            if campo_padrao in dados:
                continue
            for variante in variantes:
                for col_lower, col_real in colunas_lower.items():
                    if variante in col_lower:
                        valores = df[col_real].dropna()
                        if not valores.empty:
                            # Pegar primeiro valor não-vazio que não seja descrição
                            for val in valores:
                                val_str = str(val).strip()
                                if (val_str and len(val_str) < 100 
                                    and not val_str.lower().startswith(('para ', 'campo', 'descrição', 'número', 'endereço'))
                                    and 'obrigatório' not in val_str.lower()
                                    and 'recomendado' not in val_str.lower()):
                                    dados[campo_padrao] = val_str
                                    break
                        break
                if campo_padrao in dados:
                    break
        
        # Buscar em formato chave-valor (coluna A = campo, coluna B = valor)
        if len(df.columns) >= 2:
            for idx, row in df.iterrows():
                col0 = str(row.iloc[0]).lower().strip() if pd.notna(row.iloc[0]) else ''
                col1 = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
                
                if not col0 or not col1:
                    continue
                
                for campo_padrao, variantes in CAMPOS_BANCARIOS.items():
                    if campo_padrao in dados:
                        continue
                    for variante in variantes:
                        if variante in col0 and len(col1) < 200:
                            dados[campo_padrao] = col1
                            break
    
    return dados if dados else None


# ============================================================
# DETECÇÃO DE PLANO DE CONTAS
# ============================================================

def detectar_plano_contas(df_entrada: pd.DataFrame, todas_abas: dict = None) -> tuple:
    """
    Detecta plano de contas na planilha.
    
    Returns:
        (DataFrame com plano de contas formatado, bool se está no padrão esperado)
    """
    # Campos que indicam plano de contas
    campos_plano = ['plano de contas', 'conta contábil', 'conta financeira',
                    'categoria financeira', 'nível 1', 'nivel 1', 'classificação']
    
    fontes = {'principal': df_entrada}
    if todas_abas:
        fontes.update(todas_abas)
    
    df_plano = None
    padrao_ok = True
    
    # Buscar primeiro em abas com nome explícito de plano de contas
    fontes_ordenadas = {}
    if todas_abas:
        # Priorizar abas com "plano" ou "conta" no nome
        for nome, df in todas_abas.items():
            nome_lower = nome.lower().strip()
            if 'plano' in nome_lower or ('conta' in nome_lower and 'contato' not in nome_lower):
                fontes_ordenadas[nome] = df
        # Depois as demais
        for nome, df in todas_abas.items():
            if nome not in fontes_ordenadas:
                fontes_ordenadas[nome] = df
    fontes_ordenadas['principal'] = df_entrada
    
    for nome_fonte, df in fontes_ordenadas.items():
        if df is None or df.empty:
            continue
        
        # Verificar se a aba tem nome relacionado a plano de contas
        nome_lower = nome_fonte.lower().strip()
        eh_aba_plano = 'plano' in nome_lower or ('conta' in nome_lower and 'contato' not in nome_lower)
        
        # Pular abas que claramente não são plano de contas
        if not eh_aba_plano:
            if any(p in nome_lower for p in ['contato', 'projeto', 'modelo financ',
                                              'instrução', 'instruc', 'usuário', 'usuario',
                                              'custo', 'cobrança', 'cobranc',
                                              'atividade', 'opções', 'opcoes', 'principal',
                                              'info']):
                continue
        
        # Verificar colunas
        colunas_lower = [str(c).lower().strip() for c in df.columns]
        tem_campo_plano = any(
            any(cp in col for cp in campos_plano)
            for col in colunas_lower
        )
        
        if eh_aba_plano or tem_campo_plano:
            # Tentar extrair estrutura hierárquica (Nível 1, Nível 2, Nível 3)
            registros = []
            
            # Verificar se tem colunas de nível
            col_nivel1 = None
            col_nivel2 = None
            col_nivel3 = None
            
            for i, col in enumerate(df.columns):
                col_str = str(col).lower().strip()
                if any(x in col_str for x in ['nível 1', 'nivel 1', 'nivel 1 de classif']):
                    col_nivel1 = col
                elif any(x in col_str for x in ['nível 2', 'nivel 2', 'nivel 2 de classif']):
                    col_nivel2 = col
                elif any(x in col_str for x in ['nível 3', 'nivel 3', 'nivel 3 de classif', 'tarefa']):
                    col_nivel3 = col
            
            if col_nivel1:
                # Formato hierárquico detectado
                nivel1_atual = ''
                nivel2_atual = ''
                
                for _, row in df.iterrows():
                    n1 = str(row.get(col_nivel1, '')).strip() if pd.notna(row.get(col_nivel1)) else ''
                    n2 = str(row.get(col_nivel2, '')).strip() if col_nivel2 and pd.notna(row.get(col_nivel2)) else ''
                    n3 = str(row.get(col_nivel3, '')).strip() if col_nivel3 and pd.notna(row.get(col_nivel3)) else ''
                    
                    if n1 and n1.lower() not in ('nan', 'none', ''):
                        nivel1_atual = n1
                    if n2 and n2.lower() not in ('nan', 'none', ''):
                        nivel2_atual = n2
                    
                    # Determinar tipo
                    tipo = ''
                    if nivel1_atual.lower() in ('receitas', 'receita'):
                        tipo = 'Receita'
                    elif nivel1_atual.lower() in ('despesas fixas', 'despesas variáveis', 'despesas variaveis', 'custos', 'mão de obra', 'mao de obra'):
                        tipo = 'Despesa'
                    
                    if n3 and n3.lower() not in ('nan', 'none', ''):
                        registros.append({
                            'Tipo': tipo,
                            '1ª Categoria': nivel1_atual,
                            '2ª Categoria': nivel2_atual,
                            '3ª Categoria': n3,
                        })
                    elif n2 and not n3:
                        registros.append({
                            'Tipo': tipo,
                            '1ª Categoria': nivel1_atual,
                            '2ª Categoria': n2,
                            '3ª Categoria': '',
                        })
                
                if registros:
                    df_plano = pd.DataFrame(registros)
                    # Remover linhas vazias
                    df_plano = df_plano[df_plano['2ª Categoria'].str.strip() != ''].reset_index(drop=True)
                    padrao_ok = True
                    break
            else:
                # Formato não padrão — retornar como está para revisão
                df_plano = df.copy()
                padrao_ok = False
                break
    
    return df_plano, padrao_ok


# ============================================================
# GERAÇÃO DE PENDÊNCIAS
# ============================================================

def gerar_pendencias(df_saida: pd.DataFrame, alertas: list, plano_ok: bool, 
                     campos_nao_mapeados: list = None) -> pd.DataFrame:
    """
    Consolida todas as pendências que exigem atenção manual.
    """
    pendencias = []
    
    # Alertas de padronização
    for alerta in alertas:
        # Limpar markdown
        alerta_limpo = re.sub(r'\*\*([^*]+)\*\*', r'\1', alerta)
        pendencias.append({
            'Tipo': 'Padronização',
            'Descrição': alerta_limpo,
            'Ação Necessária': 'Cadastrar no DOit antes de importar',
        })
    
    # Plano de contas fora do padrão
    if not plano_ok:
        pendencias.append({
            'Tipo': 'Plano de Contas',
            'Descrição': 'O plano de contas foi identificado, porém sua estrutura não corresponde ao padrão esperado.',
            'Ação Necessária': 'Revisão manual necessária',
        })
    
    # Campos não mapeados
    if campos_nao_mapeados:
        for campo in campos_nao_mapeados:
            pendencias.append({
                'Tipo': 'Campo não mapeado',
                'Descrição': f'Coluna "{campo}" da origem não foi mapeada para nenhum campo padrão',
                'Ação Necessária': 'Verificar se contém dados relevantes',
            })
    
    # Campos vazios obrigatórios
    campos_obrigatorios = {
        'contatos': ['NOME'],
        'projetos': ['NOME', 'CATEGORIA'],
        'financeiro': ['DESCRIÇÃO', 'VALOR', 'VENCIMENTO'],
        'horas': ['DATA INICIAL', 'HORAS TRABALHADAS'],
    }
    
    if not df_saida.empty:
        for col in df_saida.columns:
            vazios = df_saida[col].apply(lambda x: pd.isna(x) or str(x).strip() == '').sum()
            if vazios == len(df_saida) and col in ['NOME', 'DESCRIÇÃO', 'VALOR']:
                pendencias.append({
                    'Tipo': 'Campo vazio',
                    'Descrição': f'Coluna "{col}" está completamente vazia',
                    'Ação Necessária': 'Preencher manualmente ou verificar mapeamento',
                })
    
    if not pendencias:
        return pd.DataFrame()
    
    return pd.DataFrame(pendencias)


# ============================================================
# FORMATAÇÃO DO EXCEL
# ============================================================

def _formatar_aba(ws, df):
    """Aplica formatação profissional a uma aba do Excel."""
    # Estilo do cabeçalho
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Identificar colunas de data
    colunas_data = set()
    for col_idx, col_name in enumerate(df.columns, 1):
        col_upper = str(col_name).upper()
        if any(kw in col_upper for kw in ['DATA', 'INÍCIO', 'INICIO', 'TÉRMINO', 'TERMINO', 'EXECUÇÃO', 'EXECUCAO', 'EMISSÃO', 'EMISSAO', 'VENCIMENTO']):
            colunas_data.add(col_idx)
    
    # Formatar cabeçalho
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Formatar colunas de data com formato DD/MM/YYYY
    for col_idx in colunas_data:
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != '':
                cell.number_format = 'DD/MM/YYYY'
            else:
                # Garantir que células vazias de data fiquem realmente vazias (sem tipo texto)
                cell.value = None
                cell.data_type = 'n'
    
    # Ajustar largura das colunas
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(df.columns[col_idx - 1]))
        
        for row_idx in range(2, min(len(df) + 2, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)
    
    # Adicionar filtros
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


def _criar_aba_dados_bancarios(wb, dados_bancarios: dict):
    """Cria aba de Dados Bancários no formato chave-valor."""
    ws = wb.create_sheet("Dados Bancários")
    
    # Cabeçalho
    ws.cell(row=1, column=1, value='Campo')
    ws.cell(row=1, column=2, value='Valor')
    
    # Mapa de nomes amigáveis
    nomes = {
        'favorecido': 'Favorecido',
        'cnpj': 'CNPJ/CPF',
        'banco': 'Banco',
        'agencia': 'Agência',
        'conta': 'Conta',
        'tipo_conta': 'Tipo de Conta',
        'pix': 'Chave PIX',
    }
    
    row = 2
    for campo, valor in dados_bancarios.items():
        ws.cell(row=row, column=1, value=nomes.get(campo, campo))
        ws.cell(row=row, column=2, value=valor)
        row += 1
    
    # Formatar
    df_temp = pd.DataFrame({'Campo': list(nomes.values())[:len(dados_bancarios)], 
                           'Valor': list(dados_bancarios.values())})
    _formatar_aba(ws, df_temp)
    
    # Larguras fixas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40


def _criar_aba_plano_contas(wb, df_plano: pd.DataFrame, padrao_ok: bool):
    """Cria aba de Plano de Contas."""
    ws = wb.create_sheet("Plano de Contas")
    
    # Escrever dados
    for col_idx, col_name in enumerate(df_plano.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, (_, row) in enumerate(df_plano.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else '')
    
    _formatar_aba(ws, df_plano)
    
    # Se não está no padrão, adicionar aviso
    if not padrao_ok:
        aviso_row = len(df_plano) + 3
        ws.cell(row=aviso_row, column=1, value="⚠️ ATENÇÃO: Estrutura não corresponde ao padrão esperado. Revisão manual necessária.")
        ws.cell(row=aviso_row, column=1).font = Font(bold=True, color='FF0000')


def _criar_aba_pendencias(wb, df_pendencias: pd.DataFrame):
    """Cria aba de Pendências."""
    if df_pendencias.empty:
        return
    
    ws = wb.create_sheet("Pendências")
    
    for col_idx, col_name in enumerate(df_pendencias.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, (_, row) in enumerate(df_pendencias.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else '')
    
    _formatar_aba(ws, df_pendencias)


# ============================================================
# FUNÇÃO PRINCIPAL
# ============================================================

def gerar_excel_completo(
    df_saida: pd.DataFrame,
    tipo: str,
    alertas: list = None,
    df_entrada: pd.DataFrame = None,
    todas_abas: dict = None,
    campos_nao_mapeados: list = None,
) -> bytes:
    """
    Gera o arquivo Excel final com todas as abas (dados + auxiliares).
    
    Args:
        df_saida: DataFrame convertido (dados para importação)
        tipo: Tipo de dado ('contatos', 'projetos', 'financeiro', 'horas')
        alertas: Lista de alertas de padronização
        df_entrada: DataFrame original de entrada
        todas_abas: Dict com todas as abas do arquivo original {nome: DataFrame}
        campos_nao_mapeados: Lista de colunas não mapeadas
    
    Returns:
        bytes do arquivo Excel
    """
    import io
    
    output = io.BytesIO()
    
    # Converter strings vazias para None em colunas de data
    # E garantir que valores de data são datetime (não strings)
    # (evita que o Excel grave como texto, causando erro no DOit)
    colunas_data_keywords = ['DATA', 'INÍCIO', 'INICIO', 'TÉRMINO', 'TERMINO', 
                             'EXECUÇÃO', 'EXECUCAO', 'EMISSÃO', 'EMISSAO', 'VENCIMENTO',
                             'COLUNA CUSTOMIZÁVEL DATA']
    for col in df_saida.columns:
        col_upper = str(col).upper()
        if any(kw in col_upper for kw in colunas_data_keywords):
            def _to_datetime_or_none(x):
                if x is None or (isinstance(x, str) and x.strip() == '') or (hasattr(pd, 'isna') and pd.isna(x)):
                    return None
                if isinstance(x, str):
                    try:
                        dt = pd.to_datetime(x, dayfirst=True, errors='coerce')
                        return dt if pd.notna(dt) else None
                    except Exception:
                        return None
                return x
            df_saida[col] = df_saida[col].apply(_to_datetime_or_none)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Aba principal: dados para importação
        nome_aba_principal = {
            'contatos': 'Cadastro',
            'contatos_relacionados': 'Contatos Relacionados',
            'projetos': 'Projetos',
            'financeiro': 'Financeiro',
            'horas': 'Horas',
        }.get(tipo, 'Dados')
        
        df_saida.to_excel(writer, index=False, sheet_name=nome_aba_principal)
        
        # Acessar workbook para adicionar abas formatadas
        wb = writer.book
        
        # Formatar aba principal
        ws_principal = wb[nome_aba_principal]
        _formatar_aba(ws_principal, df_saida)
        
        # --- Detectar e criar Dados Bancários (apenas financeiro) ---
        if tipo == 'financeiro':
            dados_bancarios = detectar_dados_bancarios(
                df_entrada if df_entrada is not None else pd.DataFrame(),
                todas_abas
            )
            if dados_bancarios:
                _criar_aba_dados_bancarios(wb, dados_bancarios)
        
        # --- Detectar e criar Plano de Contas (apenas financeiro) ---
        if tipo == 'financeiro':
            df_plano, plano_ok = detectar_plano_contas(
                df_entrada if df_entrada is not None else pd.DataFrame(),
                todas_abas
            )
            if df_plano is not None and not df_plano.empty:
                _criar_aba_plano_contas(wb, df_plano, plano_ok)
        else:
            df_plano = None
            plano_ok = True
        
        # --- Criar Pendências ---
        df_pendencias = gerar_pendencias(
            df_saida,
            alertas or [],
            plano_ok if df_plano is not None else True,
            campos_nao_mapeados
        )
        if not df_pendencias.empty:
            _criar_aba_pendencias(wb, df_pendencias)
    
    output.seek(0)
    return output.getvalue()
